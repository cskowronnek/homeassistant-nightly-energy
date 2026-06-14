#!/usr/bin/env python3
from __future__ import annotations

import argparse
import calendar
import json
import os
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from typing import Any
from urllib.parse import urlparse, urlunparse
from zoneinfo import ZoneInfo

import pandas as pd
import requests
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

try:
    from astral import LocationInfo
    from astral.sun import sun
except ImportError:
    LocationInfo = None
    sun = None


DEFAULT_STATISTIC_ID = "sensor.energiebezug_taglich"
DEFAULT_LOCATION_NAME = "Flensburg"
DEFAULT_LATITUDE = 54.7937
DEFAULT_LONGITUDE = 9.4469
DEFAULT_SUNSET_OFFSET_MINUTES = -60
DEFAULT_SUNRISE_OFFSET_MINUTES = 60


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Home-Assistant-Nachtverbrauch per REST abrufen und als Excel exportieren."
    )
    parser.add_argument("--ha-url", default=os.getenv("HA_URL"))
    parser.add_argument("--token", default=os.getenv("HA_TOKEN"))
    parser.add_argument("--statistic-id", default=DEFAULT_STATISTIC_ID)
    parser.add_argument("--months", type=int, default=3)
    parser.add_argument("--timezone", default="Europe/Berlin")
    parser.add_argument(
        "--night-mode",
        choices=("sun", "fixed"),
        default="sun",
        help="Nachtdefinition: Sonnenzeiten oder fixes Stundenfenster.",
    )
    parser.add_argument("--night-start", type=int, default=18)
    parser.add_argument("--night-end", type=int, default=7)
    parser.add_argument("--location-name", default=DEFAULT_LOCATION_NAME)
    parser.add_argument("--latitude", type=float, default=DEFAULT_LATITUDE)
    parser.add_argument("--longitude", type=float, default=DEFAULT_LONGITUDE)
    parser.add_argument(
        "--sunset-offset-minutes",
        type=int,
        default=DEFAULT_SUNSET_OFFSET_MINUTES,
    )
    parser.add_argument(
        "--sunrise-offset-minutes",
        type=int,
        default=DEFAULT_SUNRISE_OFFSET_MINUTES,
    )
    parser.add_argument("--output", default="nachtverbrauch_homeassistant.xlsx")
    parser.add_argument(
        "--verify-ssl",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="SSL-Zertifikat prüfen. Mit --no-verify-ssl deaktivieren.",
    )
    return parser.parse_args()


def normalize_ha_base_url(ha_url: str) -> str:
    if "://" not in ha_url:
        ha_url = "http://" + ha_url

    parsed = urlparse(ha_url)
    path = parsed.path.rstrip("/")

    if path.endswith("/api"):
        path = path[:-4]

    normalized = parsed._replace(path=path, params="", query="", fragment="")
    return urlunparse(normalized).rstrip("/")


def subtract_months(dt: datetime, months: int) -> datetime:
    month_index = dt.month - 1 - months
    year = dt.year + month_index // 12
    month = month_index % 12 + 1
    day = min(dt.day, calendar.monthrange(year, month)[1])
    return dt.replace(year=year, month=month, day=day)


def iso_utc(dt: datetime) -> str:
    return (
        dt.astimezone(timezone.utc)
        .isoformat(timespec="milliseconds")
        .replace("+00:00", "Z")
    )


def parse_ha_timestamp(value: Any) -> datetime:
    if isinstance(value, (int, float)):
        seconds = value / 1000 if value > 10_000_000_000 else value
        return datetime.fromtimestamp(seconds, tz=timezone.utc)

    if isinstance(value, str):
        return datetime.fromisoformat(value.replace("Z", "+00:00")).astimezone(timezone.utc)

    raise TypeError(f"Unbekannter Zeitstempeltyp: {value!r}")


def rest_request_statistics(
    ha_url: str,
    token: str,
    statistic_id: str,
    start_utc: datetime,
    end_utc: datetime,
    verify_ssl: bool = True,
) -> dict[str, list[dict[str, Any]]]:
    base_url = normalize_ha_base_url(ha_url)
    url = f"{base_url}/api/services/recorder/get_statistics?return_response"

    headers = {
        "Authorization": f"Bearer {token}",
        "Content-Type": "application/json",
    }

    payload = {
        "statistic_ids": [statistic_id],
        "start_time": iso_utc(start_utc),
        "end_time": iso_utc(end_utc),
        "period": "hour",
        "types": ["change"],
        "units": {
            "energy": "kWh",
        },
    }

    response = requests.post(
        url,
        headers=headers,
        json=payload,
        timeout=120,
        verify=verify_ssl,
    )

    try:
        response.raise_for_status()
    except requests.HTTPError as exc:
        body = response.text[:2000]
        raise RuntimeError(
            f"Home-Assistant-REST-Aufruf fehlgeschlagen: HTTP {response.status_code}\n"
            f"URL: {url}\n"
            f"Antwort: {body}\n\n"
            "Hinweise:\n"
            "- Benötigt Home Assistant 2025.6 oder neuer für recorder.get_statistics.\n"
            "- Der Token braucht Zugriff auf Recorder/Services.\n"
            "- Bei selbstsigniertem HTTPS-Zertifikat: --no-verify-ssl nutzen."
        ) from exc

    body = response.json()
    service_response = body.get("service_response", {})

    if "statistics" in service_response:
        return service_response["statistics"]

    if statistic_id in service_response:
        return service_response

    raise RuntimeError(
        "Unerwartete Antwortstruktur von recorder.get_statistics:\n"
        + json.dumps(body, indent=2, ensure_ascii=False)[:3000]
    )


def require_astral() -> None:
    if LocationInfo is None or sun is None:
        raise RuntimeError(
            "Das Paket 'astral' fehlt. Installiere es mit: python -m pip install astral"
        )


def elapsed_seconds(start: datetime, end: datetime) -> float:
    return (
        end.astimezone(timezone.utc) - start.astimezone(timezone.utc)
    ).total_seconds()


def overlap_seconds(
    first_start: datetime,
    first_end: datetime,
    second_start: datetime,
    second_end: datetime,
) -> float:
    latest_start = max(
        first_start.astimezone(timezone.utc),
        second_start.astimezone(timezone.utc),
    )
    earliest_end = min(
        first_end.astimezone(timezone.utc),
        second_end.astimezone(timezone.utc),
    )
    return max(0.0, (earliest_end - latest_start).total_seconds())


def fixed_window_for_night(
    night: date,
    tz: ZoneInfo,
    night_start: int,
    night_end: int,
) -> dict[str, Any]:
    start = datetime.combine(night, time(night_start), tzinfo=tz)
    end_day = night + timedelta(days=1 if night_start >= night_end else 0)
    end = datetime.combine(end_day, time(night_end), tzinfo=tz)

    return {
        "nacht": night,
        "nacht_start_lokal": start,
        "sonnenuntergang_lokal": pd.NaT,
        "sonnenaufgang_lokal": pd.NaT,
        "nacht_ende_lokal": end,
        "nachtfenster_minuten": elapsed_seconds(start, end) / 60,
    }


def sun_window_for_night(
    night: date,
    tz_name: str,
    tz: ZoneInfo,
    location_name: str,
    latitude: float,
    longitude: float,
    sunset_offset_minutes: int,
    sunrise_offset_minutes: int,
) -> dict[str, Any]:
    require_astral()

    location = LocationInfo(location_name, "Germany", tz_name, latitude, longitude)

    try:
        sunset = sun(location.observer, date=night, tzinfo=tz)["sunset"]
        sunrise = sun(
            location.observer,
            date=night + timedelta(days=1),
            tzinfo=tz,
        )["sunrise"]
    except Exception as exc:
        raise RuntimeError(
            f"Sonnenzeiten fuer {location_name} am {night} konnten nicht berechnet werden."
        ) from exc

    start = sunset + timedelta(minutes=sunset_offset_minutes)
    end = sunrise + timedelta(minutes=sunrise_offset_minutes)

    return {
        "nacht": night,
        "nacht_start_lokal": start,
        "sonnenuntergang_lokal": sunset,
        "sonnenaufgang_lokal": sunrise,
        "nacht_ende_lokal": end,
        "nachtfenster_minuten": elapsed_seconds(start, end) / 60,
    }


def build_night_windows(
    start_date: date,
    end_date: date,
    tz_name: str,
    tz: ZoneInfo,
    night_mode: str,
    night_start: int,
    night_end: int,
    location_name: str,
    latitude: float,
    longitude: float,
    sunset_offset_minutes: int,
    sunrise_offset_minutes: int,
) -> list[dict[str, Any]]:
    windows = []
    current = start_date

    while current <= end_date:
        if night_mode == "sun":
            windows.append(
                sun_window_for_night(
                    current,
                    tz_name,
                    tz,
                    location_name,
                    latitude,
                    longitude,
                    sunset_offset_minutes,
                    sunrise_offset_minutes,
                )
            )
        elif night_mode == "fixed":
            windows.append(fixed_window_for_night(current, tz, night_start, night_end))
        else:
            raise ValueError(f"Unbekannter night-mode: {night_mode!r}")

        current += timedelta(days=1)

    return windows


def build_dataframes(
    raw_result: dict[str, list[dict[str, Any]]],
    statistic_id: str,
    tz_name: str,
    night_mode: str,
    night_start: int,
    night_end: int,
    location_name: str,
    latitude: float,
    longitude: float,
    sunset_offset_minutes: int,
    sunrise_offset_minutes: int,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    rows = raw_result.get(statistic_id, [])

    if not rows:
        raise RuntimeError(
            f"Keine Statistikdaten für {statistic_id!r} erhalten.\n"
            "Prüfe in Home Assistant unter Entwicklerwerkzeuge → Statistik, ob diese Entität "
            "Long-Term-Statistics erzeugt und ob 'change' verfügbar ist."
        )

    records = []

    for row in rows:
        if "start" not in row:
            continue

        end_value = row.get("end")

        if end_value is None:
            end_dt = parse_ha_timestamp(row["start"]) + timedelta(hours=1)
        else:
            end_dt = parse_ha_timestamp(end_value)

        records.append(
            {
                "zeit_start_utc": parse_ha_timestamp(row["start"]),
                "zeit_ende_utc": end_dt,
                "verbrauch_kWh": row.get("change"),
            }
        )

    if not records:
        raise RuntimeError("Die Antwort enthielt keine verwertbaren Statistikzeilen.")

    df = pd.DataFrame.from_records(records)
    df["verbrauch_kWh"] = pd.to_numeric(df["verbrauch_kWh"], errors="coerce")

    df["zeit_start_utc"] = pd.to_datetime(df["zeit_start_utc"], utc=True)
    df["zeit_ende_utc"] = pd.to_datetime(df["zeit_ende_utc"], utc=True)

    df["zeit_start_lokal"] = df["zeit_start_utc"].dt.tz_convert(tz_name)
    df["zeit_ende_lokal"] = df["zeit_ende_utc"].dt.tz_convert(tz_name)
    tz = ZoneInfo(tz_name)

    window_start_date = df["zeit_start_lokal"].min().date() - timedelta(days=1)
    window_end_date = df["zeit_ende_lokal"].max().date()
    windows = build_night_windows(
        window_start_date,
        window_end_date,
        tz_name,
        tz,
        night_mode,
        night_start,
        night_end,
        location_name,
        latitude,
        longitude,
        sunset_offset_minutes,
        sunrise_offset_minutes,
    )

    detail_records = []

    for row in df.itertuples(index=False):
        row_start = row.zeit_start_lokal
        row_end = row.zeit_ende_lokal
        row_seconds = elapsed_seconds(row_start, row_end)

        if row_seconds <= 0:
            continue

        for window in windows:
            seconds = overlap_seconds(
                row_start,
                row_end,
                window["nacht_start_lokal"],
                window["nacht_ende_lokal"],
            )

            if seconds <= 0:
                continue

            fraction = seconds / row_seconds
            verbrauch_stunde = (
                None if pd.isna(row.verbrauch_kWh) else float(row.verbrauch_kWh)
            )
            nachtverbrauch = (
                None if verbrauch_stunde is None else verbrauch_stunde * fraction
            )

            detail_records.append(
                {
                    "nacht": window["nacht"],
                    "nacht_start_lokal": window["nacht_start_lokal"],
                    "sonnenuntergang_lokal": window["sonnenuntergang_lokal"],
                    "sonnenaufgang_lokal": window["sonnenaufgang_lokal"],
                    "nacht_ende_lokal": window["nacht_ende_lokal"],
                    "nachtfenster_minuten": window["nachtfenster_minuten"],
                    "zeit_start_lokal": row_start,
                    "zeit_ende_lokal": row_end,
                    "stunde": f"{row_start:%H:%M}-{row_end:%H:%M}",
                    "verbrauch_stunde_kWh": verbrauch_stunde,
                    "ueberlappung_minuten": seconds / 60,
                    "anteil_im_nachtfenster": fraction,
                    "nachtverbrauch_anteilig_kWh": nachtverbrauch,
                    "statistic_id": statistic_id,
                    "zeit_start_utc": row.zeit_start_utc,
                    "zeit_ende_utc": row.zeit_ende_utc,
                }
            )

    if not detail_records:
        raise RuntimeError(
            "Es wurden Daten gefunden, aber keine Ueberschneidung mit dem Nachtfenster."
        )

    details = pd.DataFrame.from_records(detail_records)

    numeric_columns = [
        "nachtfenster_minuten",
        "verbrauch_stunde_kWh",
        "ueberlappung_minuten",
        "anteil_im_nachtfenster",
        "nachtverbrauch_anteilig_kWh",
    ]
    for col in numeric_columns:
        details[col] = pd.to_numeric(details[col], errors="coerce")

    details["minuten_mit_daten"] = details["ueberlappung_minuten"].where(
        details["verbrauch_stunde_kWh"].notna(),
        0.0,
    )

    details = details[
        [
            "nacht",
            "nacht_start_lokal",
            "sonnenuntergang_lokal",
            "sonnenaufgang_lokal",
            "nacht_ende_lokal",
            "nachtfenster_minuten",
            "zeit_start_lokal",
            "zeit_ende_lokal",
            "stunde",
            "verbrauch_stunde_kWh",
            "ueberlappung_minuten",
            "minuten_mit_daten",
            "anteil_im_nachtfenster",
            "nachtverbrauch_anteilig_kWh",
            "statistic_id",
            "zeit_start_utc",
            "zeit_ende_utc",
        ]
    ].sort_values(["nacht", "zeit_start_lokal"])

    nightly = details.groupby("nacht", as_index=False).agg(
        nacht_start_lokal=("nacht_start_lokal", "first"),
        sonnenuntergang_lokal=("sonnenuntergang_lokal", "first"),
        sonnenaufgang_lokal=("sonnenaufgang_lokal", "first"),
        nacht_ende_lokal=("nacht_ende_lokal", "first"),
        nachtfenster_minuten=("nachtfenster_minuten", "first"),
        nachtverbrauch_kWh=("nachtverbrauch_anteilig_kWh", "sum"),
        minuten_mit_daten=("minuten_mit_daten", "sum"),
        negative_stunden=("verbrauch_stunde_kWh", lambda s: int((s < 0).sum())),
    )

    nightly["minuten_erwartet"] = nightly["nachtfenster_minuten"]
    nightly["stunden_mit_daten"] = nightly["minuten_mit_daten"] / 60
    nightly["stunden_erwartet"] = nightly["minuten_erwartet"] / 60
    nightly["vollstaendig"] = (
        nightly["minuten_mit_daten"] + 0.5 >= nightly["minuten_erwartet"]
    )

    nightly = nightly[
        [
            "nacht",
            "nacht_start_lokal",
            "sonnenuntergang_lokal",
            "sonnenaufgang_lokal",
            "nacht_ende_lokal",
            "nachtverbrauch_kWh",
            "minuten_mit_daten",
            "minuten_erwartet",
            "stunden_mit_daten",
            "stunden_erwartet",
            "vollstaendig",
            "negative_stunden",
        ]
    ]

    monthly = nightly.copy()
    monthly["monat"] = pd.to_datetime(monthly["nacht"]).dt.to_period("M").astype(str)

    monthly = monthly.groupby("monat", as_index=False).agg(
        nachtverbrauch_kWh=("nachtverbrauch_kWh", "sum"),
        naechte=("nacht", "count"),
        vollstaendige_naechte=("vollstaendig", "sum"),
        durchschnitt_pro_nacht_kWh=("nachtverbrauch_kWh", "mean"),
        max_nacht_kWh=("nachtverbrauch_kWh", "max"),
        min_nacht_kWh=("nachtverbrauch_kWh", "min"),
    )

    return details, nightly, monthly


def make_excel_safe(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    for col in out.columns:
        if isinstance(out[col].dtype, pd.DatetimeTZDtype):
            out[col] = out[col].dt.tz_localize(None)
        elif out[col].dtype == "object":
            out[col] = out[col].map(make_excel_safe_value)

    return out


def make_excel_safe_value(value: Any) -> Any:
    if value is None or pd.isna(value):
        return None

    if isinstance(value, pd.Timestamp):
        if value.tzinfo is not None:
            return value.tz_localize(None).to_pydatetime()
        return value.to_pydatetime()

    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)

    return value


def format_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for column_cells in ws.columns:
        max_len = max(
            len(str(cell.value)) if cell.value is not None else 0
            for cell in column_cells
        )
        col_letter = get_column_letter(column_cells[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 42)


def find_header_col(ws, header: str) -> int:
    for cell in ws[1]:
        if cell.value == header:
            return cell.column

    raise RuntimeError(f"Spalte {header!r} wurde im Blatt {ws.title!r} nicht gefunden.")


def write_excel(
    output: str,
    details: pd.DataFrame,
    nightly: pd.DataFrame,
    monthly: pd.DataFrame,
    overview: pd.DataFrame,
) -> None:
    output_path = Path(output)

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
        datetime_format="yyyy-mm-dd hh:mm:ss",
    ) as writer:
        make_excel_safe(overview).to_excel(writer, sheet_name="Überblick", index=False)
        make_excel_safe(monthly).to_excel(writer, sheet_name="Monate", index=False)
        make_excel_safe(nightly).to_excel(writer, sheet_name="Nächte", index=False)
        make_excel_safe(details).to_excel(writer, sheet_name="Stundenwerte", index=False)

        wb = writer.book

        for ws in wb.worksheets:
            format_sheet(ws)

        ws = wb["Nächte"]

        if ws.max_row > 1:
            chart = BarChart()
            chart.title = "Nachtverbrauch pro Nacht"
            chart.y_axis.title = "kWh"
            chart.x_axis.title = "Nacht"

            value_col = find_header_col(ws, "nachtverbrauch_kWh")
            data = Reference(ws, min_col=value_col, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 8
            chart.width = 20

            chart_anchor = f"{get_column_letter(ws.max_column + 2)}2"
            ws.add_chart(chart, chart_anchor)


def query_start_for_args(args: argparse.Namespace, start_day: date, tz: ZoneInfo) -> datetime:
    if args.night_mode == "sun":
        window = sun_window_for_night(
            start_day,
            args.timezone,
            tz,
            args.location_name,
            args.latitude,
            args.longitude,
            args.sunset_offset_minutes,
            args.sunrise_offset_minutes,
        )
        return window["nacht_start_lokal"]

    return datetime.combine(start_day, time(args.night_start), tzinfo=tz)


def night_definition_for_args(args: argparse.Namespace) -> str:
    if args.night_mode == "sun":
        return (
            "Sonnenuntergang "
            f"{args.sunset_offset_minutes:+d} Min. bis Sonnenaufgang "
            f"{args.sunrise_offset_minutes:+d} Min."
        )

    return f"{args.night_start:02d}:00-{args.night_end:02d}:00"


def main() -> None:
    args = parse_args()

    if not args.ha_url:
        raise SystemExit("Bitte --ha-url angeben oder Umgebungsvariable HA_URL setzen.")

    if not args.token:
        raise SystemExit("Bitte --token angeben oder Umgebungsvariable HA_TOKEN setzen.")

    tz = ZoneInfo(args.timezone)
    now_local = datetime.now(tz)
    start_day = subtract_months(now_local, args.months).date()

    try:
        query_start_local = query_start_for_args(args, start_day, tz)
    except RuntimeError as exc:
        raise SystemExit(str(exc)) from exc

    query_end_local = now_local

    raw = rest_request_statistics(
        ha_url=args.ha_url,
        token=args.token,
        statistic_id=args.statistic_id,
        start_utc=query_start_local.astimezone(timezone.utc),
        end_utc=query_end_local.astimezone(timezone.utc),
        verify_ssl=args.verify_ssl,
    )

    details, nightly, monthly = build_dataframes(
        raw,
        statistic_id=args.statistic_id,
        tz_name=args.timezone,
        night_mode=args.night_mode,
        night_start=args.night_start,
        night_end=args.night_end,
        location_name=args.location_name,
        latitude=args.latitude,
        longitude=args.longitude,
        sunset_offset_minutes=args.sunset_offset_minutes,
        sunrise_offset_minutes=args.sunrise_offset_minutes,
    )

    overview = pd.DataFrame(
        [
            {
                "statistic_id": args.statistic_id,
                "abfrage_start_lokal": query_start_local,
                "abfrage_ende_lokal": query_end_local,
                "nachtmodus": args.night_mode,
                "nachtfenster": night_definition_for_args(args),
                "zeitzone": args.timezone,
                "ort": args.location_name if args.night_mode == "sun" else None,
                "latitude": args.latitude if args.night_mode == "sun" else None,
                "longitude": args.longitude if args.night_mode == "sun" else None,
                "sunset_offset_minutes": (
                    args.sunset_offset_minutes if args.night_mode == "sun" else None
                ),
                "sunrise_offset_minutes": (
                    args.sunrise_offset_minutes if args.night_mode == "sun" else None
                ),
                "fixed_night_start": (
                    args.night_start if args.night_mode == "fixed" else None
                ),
                "fixed_night_end": args.night_end if args.night_mode == "fixed" else None,
                "gesamt_nachtverbrauch_kWh": nightly["nachtverbrauch_kWh"].sum(),
                "anzahl_naechte": len(nightly),
                "vollstaendige_naechte": int(nightly["vollstaendig"].sum()),
                "durchschnitt_kWh_pro_nacht": nightly["nachtverbrauch_kWh"].mean(),
                "datei_erzeugt_am": now_local,
            }
        ]
    )

    write_excel(args.output, details, nightly, monthly, overview)

    print(f"Fertig: {args.output}")
    print(f"Nächte: {len(nightly)}")
    print(f"Gesamt: {nightly['nachtverbrauch_kWh'].sum():.3f} kWh")
    print("Randstunden wurden anteilig nach Ueberschneidung mit dem Nachtfenster gewichtet.")
    print("Unvollständige Nächte sind im Blatt 'Nächte' markiert.")
    print("Negative Stundenwerte werden im Blatt 'Nächte' gezählt.")


if __name__ == "__main__":
    main()
